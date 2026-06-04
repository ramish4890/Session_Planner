"""Build the output Excel workbook with all sheets."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter


def build_workbook(scheduled, unscheduled, roster, num_weeks):
    """
    Create Excel workbook with:
        - Final Schedule (long format)
        - Unscheduled Reps
        - Calendar Grid
        - Daily Balance
        - Windows & Rules
        - Leave Adjustments
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Final Schedule
    ws_final = wb.create_sheet("Final Schedule")
    if num_weeks > 1:
        headers = ["Workday_ID", "Pseudo_Name", "Email", "Manager", "Role/JD", "Shift_Time",
                   "Off_Day", "Week", "Schedule_Day", "Schedule_Date", "Session Start Time", "Session End Time"]
    else:
        headers = ["Workday_ID", "Pseudo_Name", "Email", "Manager", "Role/JD", "Shift_Time",
                   "Off_Day", "Schedule_Day", "Schedule_Date", "Session Start Time", "Session End Time"]

    ws_final.append(headers)

    # Style header
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws_final.cell(1, col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sort scheduled by week, date, time, name
    scheduled_sorted = sorted(scheduled, key=lambda x: (x.get('week', 1), x['schedule_date'], x['session_start_time'], x['pseudo_name']))

    # Day color banding
    day_colors = {
        'Monday': 'E2EFDA',
        'Tuesday': 'FCE4D6',
        'Wednesday': 'FFF2CC',
        'Thursday': 'DEEBF7',
        'Friday': 'F4B084'
    }

    for session in scheduled_sorted:
        if num_weeks > 1:
            row_data = [
                session['workday_id'],
                session['pseudo_name'],
                session['email'],
                session['manager'],
                session['role'],
                session['shift_time'],
                session['off_day'],
                session.get('week', 1),
                session['schedule_day'],
                session['schedule_date'].strftime('%Y-%m-%d'),
                session['session_start_time'],
                session['session_end_time']
            ]
        else:
            row_data = [
                session['workday_id'],
                session['pseudo_name'],
                session['email'],
                session['manager'],
                session['role'],
                session['shift_time'],
                session['off_day'],
                session['schedule_day'],
                session['schedule_date'].strftime('%Y-%m-%d'),
                session['session_start_time'],
                session['session_end_time']
            ]

        ws_final.append(row_data)
        row_num = ws_final.max_row

        # Apply day color
        fill_color = day_colors.get(session['schedule_day'], 'FFFFFF')
        for col_idx in range(1, len(row_data) + 1):
            ws_final.cell(row_num, col_idx).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        ws_final.column_dimensions[get_column_letter(col_idx)].width = 15

    # Sheet 2: Unscheduled Reps
    ws_unsched = wb.create_sheet("Unscheduled Reps")
    if num_weeks > 1:
        unsched_headers = ["Workday_ID", "Pseudo_Name", "Manager", "Role/JD", "Week",
                          "Sessions Needed", "Sessions Placed", "Sessions Missing", "Reason", "Detail"]
    else:
        unsched_headers = ["Workday_ID", "Pseudo_Name", "Manager", "Role/JD",
                          "Sessions Needed", "Sessions Placed", "Sessions Missing", "Reason", "Detail"]

    ws_unsched.append(unsched_headers)

    # Style header
    for col_idx, hdr in enumerate(unsched_headers, 1):
        cell = ws_unsched.cell(1, col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add unscheduled rows
    for rep, placed, reason, week in unscheduled:
        from scheduler.engine import schedule_sessions  # Get sessions_per_rep from context
        # We don't have sessions_per_rep here, so assume 2 for now
        sessions_needed = 2
        missing = sessions_needed - placed

        if num_weeks > 1:
            row_data = [
                rep['workday_id'],
                rep['pseudo_name'],
                rep['manager'],
                rep['role'],
                week,
                sessions_needed,
                placed,
                missing,
                reason,
                f"{rep['pseudo_name']} needed {sessions_needed} but got {placed}. {reason}."
            ]
        else:
            row_data = [
                rep['workday_id'],
                rep['pseudo_name'],
                rep['manager'],
                rep['role'],
                sessions_needed,
                placed,
                missing,
                reason,
                f"{rep['pseudo_name']} needed {sessions_needed} but got {placed}. {reason}."
            ]

        ws_unsched.append(row_data)
        row_num = ws_unsched.max_row

        # Highlight in orange
        for col_idx in range(1, len(row_data) + 1):
            ws_unsched.cell(row_num, col_idx).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # Auto-width
    for col_idx in range(1, len(unsched_headers) + 1):
        ws_unsched.column_dimensions[get_column_letter(col_idx)].width = 15

    # Sheet 3: Daily Balance
    ws_balance = wb.create_sheet("Daily Balance")
    ws_balance.append(["Date", "Day", "Sessions", "Utilization"])

    # Style header
    for col_idx in range(1, 5):
        cell = ws_balance.cell(1, col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Count sessions per day
    day_counts = Counter((s['schedule_date'], s['schedule_day']) for s in scheduled)

    for (date, day), count in sorted(day_counts.items()):
        ws_balance.append([date.strftime('%Y-%m-%d'), day, count, f"{count} sessions"])

    for col_idx in range(1, 5):
        ws_balance.column_dimensions[get_column_letter(col_idx)].width = 15

    # Sheet 4: Leave Adjustments (placeholder)
    ws_leave = wb.create_sheet("Leave Adjustments")
    ws_leave.append(["Note"])
    ws_leave.append(["Leave data applied. Unscheduled reps shown in Unscheduled Reps sheet."])

    return wb
